# Reads data from ./temp/export.xlsx or export.csv: row 3 = headers, then data. Maps via header_map, upserts by registry_number.

import csv
import warnings
import re
from datetime import datetime as dt
from pathlib import Path
from typing import Any, Optional

from models.inagents import AGENT_TYPE_MAP

with warnings.catch_warnings():
    warnings.simplefilter("ignore", UserWarning)
    from openpyxl import load_workbook

from extensions import db
from models import Inagent

MAX_HEADER_COLS = 200
HEADER_ROW = 3
DATA_START_ROW = 4

DATE_COLUMNS = frozenset({
    "birth_date", "include_minjust_date", "exclude_minjust_date",
    "publish_date", "account_open_date", "contract_date",
})

HEADER_MAP: dict[str, str] = {
    "№ п/п": "registry_number",
    "Полное наименование (прежнее наименование (в случае его изменения)) / ФИО «Псевдоним» (при наличии) (прежние ФИО (в случае их изменения))": "full_name",
    "Основания для включения": "include_reason",
    "Дата принятия Минюстом России решения о включении в реестр": "include_minjust_date",
    "Дата принятия Минюстом России решения об исключении из реестра (при наличии)": "exclude_minjust_date",
    "Доменное имя информационного ресурса (при наличии)": "domain_name",
    "Тип иностранного агента": "agent_type",
    "Регистрационный номер": "reg_num",
    "ИНН": "inn",
    "ОГРН": "ogrn",
    "СНИЛС": "snils",
    "Дата рождения": "birth_date",
    "Полное наименование или ФИО участников": "participants",
    "Адрес (место нахождения)": "address",
    "Дата опубликования принятого Минюстом России решения о включении в реестр": "publish_date",
    "Номер специального счета": "special_account_num",
    "Наименование и местонахождение уполномоченного банка": "bank_name_location",
    "Банковский идентификационный код уполномоченного банка": "bank_bik",
    "Номер корреспондентского счета (субсчета) уполномоченного банка (его филиала)": "bank_corr_account",
    "Дата открытия специального счета": "account_open_date",
    "Дата заключения договора банковского счёта": "contract_date",
}

DB_COLUMNS_FROM_FILE = set(HEADER_MAP.values())


def _parse_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_agent_type(raw_agent_type: str = "") -> Optional[str]:
    raw_agent_type = str(raw_agent_type).strip()

    for key, raw_value in AGENT_TYPE_MAP.items():
        if raw_value.lower() == raw_agent_type.lower():
            return key

    return None


def _normalize_value(col: str, value: Any) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if col == "registry_number":
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None
    if col in DATE_COLUMNS:
        return _parse_date(value)
    if col == "agent_type":
        return _normalize_agent_type(value)
    if col == "domain_name":
        s = str(value).strip()
        if not s:
            return None
        parts = []
        for sep in (";", ","):
            if sep in s:
                parts = [x.strip() for x in s.split(sep) if x.strip()]
                break
        return parts if parts else [s]
    return str(value).strip() or None


class InagentsXlsxParser:
    """Reads xlsx: row 3 = headers, rows 4+ = data. Builds dicts with file headers, then maps to DB column names. Upserts by registry_number."""

    def __init__(self, file_path: Path) -> None:
        self.file_path = file_path

    def _raw_value_for_header(self, raw: dict[str, Any], file_header: str) -> Any:
        key_stripped = file_header.strip()
        for k, v in raw.items():
            if (k or "").strip() == key_stripped:
                return v
        return None

    def _read_headers_xlsx(self, ws: Any) -> list[str]:
        row_iter = ws.iter_rows(
            min_row=HEADER_ROW,
            max_row=HEADER_ROW,
            min_col=1,
            max_col=MAX_HEADER_COLS,
            values_only=True,
        )
        raw = [str(c).strip() if c is not None else "" for c in next(row_iter)]
        while raw and raw[-1] == "":
            raw.pop()
        return raw

    def _load_raw_rows_csv(self) -> list[dict[str, Any]]:
        with open(self.file_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter=";")
            lines = list(reader)
        if len(lines) < DATA_START_ROW:
            return []
        headers = [h.strip() for h in lines[HEADER_ROW - 1]]
        rows: list[dict[str, Any]] = []
        for row_cells in lines[DATA_START_ROW - 1:]:
            if len(row_cells) < len(headers):
                row_cells = row_cells + [""] * (len(headers) - len(row_cells))
            values = [c.strip() or None for c in row_cells[: len(headers)]]
            if all(v is None for v in values):
                continue
            rows.append(dict(zip(headers, values)))
        return rows

    def _load_raw_rows_xlsx(self) -> list[dict[str, Any]]:
        wb = load_workbook(self.file_path, read_only=False, data_only=True)
        ws = wb.active
        headers = self._read_headers_xlsx(ws)
        rows: list[dict[str, Any]] = []
        for row_tuple in ws.iter_rows(
                min_row=DATA_START_ROW,
                min_col=1,
                max_col=len(headers),
                values_only=True,
        ):
            row_list = list(row_tuple)
            if len(row_list) < len(headers):
                row_list.extend([None] * (len(headers) - len(row_list)))
            rows.append(dict(zip(headers, row_list[: len(headers)])))
        wb.close()
        return rows

    def load_raw_rows(self) -> list[dict[str, Any]]:
        """Rows as dicts with original (file) header names as keys."""
        if self.file_path.suffix.lower() == ".csv":
            return self._load_raw_rows_csv()
        return self._load_raw_rows_xlsx()

    def _map_row(self, raw: dict[str, Any]) -> dict[str, Any]:
        """Transform dict: file headers -> DB column names; normalize values."""
        mapped: dict[str, Any] = {}
        for file_header, db_col in HEADER_MAP.items():
            value = self._raw_value_for_header(raw, file_header)
            mapped[db_col] = _normalize_value(db_col, value)
        return mapped

    def load_mapped_rows(self) -> list[dict[str, Any]]:
        """Raw rows transformed to dicts with DB column names."""
        return [self._map_row(r) for r in self.load_raw_rows()]

    def sync_to_db(self, rows: Optional[list[dict[str, Any]]] = None) -> tuple[int, int]:
        """Upsert rows: by registry_number update existing or insert. Returns (inserted, updated). Requires Flask app context."""
        if rows is None:
            rows = self.load_mapped_rows()
        inserted = 0
        updated = 0

        for row in rows:
            registry_number = row.get("registry_number")

            if registry_number is None:
                continue

            inagent: Inagent = Inagent.query.filter_by(registry_number=registry_number).first()
            payload = {k: v for k, v in row.items() if k in DB_COLUMNS_FROM_FILE}

            if inagent:
                for key, value in payload.items():
                    setattr(inagent, key, value)

                if inagent.search_terms is None or len(inagent.search_terms) == 0:
                    inagent.search_terms = self._parse_search_terms(inagent.full_name)

                updated += 1
            else:
                db.session.add(Inagent(**payload))
                inserted += 1

        db.session.commit()

        return (inserted, updated)

    @staticmethod
    def _parse_fio_with_nicknames(full_name: str) -> list:
        """
        Парсит строку "Фамилия Имя Отчество \"псевдоним1, псевдоним2\""
        и возвращает список строк в нужных форматах.
        """
        # Шаблон: ФИО + опциональные псевдонимы в кавычках
        pattern = r'((?:[А-ЯЁ]{1}[а-яё]{1,}\s?){3})(?:\"(.+?)\")?'
        match = re.match(pattern, full_name.strip())

        search_phrases = []

        if not match:
            return search_phrases

        full_name, nicknames = match.groups()

        # [start] parse name
        surname, first_name, patronymic = [name for name in full_name.strip().split(" ")]
        search_phrases.append(f"{surname} {first_name} {patronymic}")
        search_phrases.append(f"{surname} {first_name[0]}. {patronymic[0]}.")
        # [end]

        # pase nick
        if nicknames is not None:
            search_phrases.extend([nick.strip() for nick in nicknames.split(',')])

        return search_phrases

    @staticmethod
    def _parse_organisation_name(full_name: str):
        pattern = r'(?:[А-яЁё]+\s)+?«(.*?)»'
        match = re.match(pattern, full_name.strip())
        search_phrases = []

        if match is None:
            return search_phrases

        org_title = match.groups()[0]

        if org_title:
            search_phrases.append(org_title)

        return search_phrases

    @staticmethod
    def _parse_search_terms(full_name: str):
        terms = []
        terms.extend(InagentsXlsxParser._parse_fio_with_nicknames(full_name))
        terms.extend(InagentsXlsxParser._parse_organisation_name(full_name))

        return list(set(terms))

def main() -> None:
    script_dir = Path(__file__).resolve().parent
    temp_dir = script_dir / "temp"
    file_path = temp_dir / "export.csv" if (temp_dir / "export.csv").exists() else temp_dir / "export.xlsx"
    parser = InagentsXlsxParser(file_path)
    rows = parser.load_mapped_rows()
    print(len(rows))
    if rows:
        print(rows[0])


if __name__ == "__main__":
    from app import create_app

    app = create_app()
    with app.app_context():
        main()
